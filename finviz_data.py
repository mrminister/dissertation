import requests
import csv
from openpyxl import load_workbook
from bs4 import BeautifulSoup

counter = 1
company_list = []
with open('russell_3000_2011-06-27.csv', 'r') as csv_file:
    reader = csv.reader(csv_file, delimiter=' ', quotechar='|')
    for row in reader:
        company_list.append(str(row).strip("'[]"))


def scrape_site(ticker_name):
    address = 'https://finviz.com/quote.ashx?t='+ticker_name
    r = requests.get(address)
    soup = BeautifulSoup(r.text)
    table = soup.find_all('table', class_='fullview-ratings-outer')
    dates = BeautifulSoup(str(table)).find_all('td', align="left", width="120")
    advisors = BeautifulSoup(str(table)).find_all('td', align="left", width="250")
    recommendations = BeautifulSoup(str(table)).find_all('td', align="left", width="150")

    #spreadsheet org and sorting
    wb = load_workbook(filename='base_data.xlsx')
    ws = wb.worksheets[0]
    for l in range(max(len(dates),len(advisors),len(recommendations))):
        try:
            date = str(dates[l].contents).strip("'[]")
            advisor = str(advisors[l].contents).strip("'[]")
            feedback = str(recommendations[l].contents).strip("'[]$").split('→')[0]
            feedback_ = str(recommendations[l].contents).strip("'[]$").split('→')[1].strip('$')
            global counter
            ticker_i = 'A'+str(counter)
            date_i = 'B'+str(counter)
            institution_i = 'C'+str(counter)
            feedback_i = 'D'+str(counter)
            feedback_after = 'E' + str(counter)

            ws[ticker_i] = ticker_name
            ws[date_i] = date
            ws[institution_i] = advisor
            ws[feedback_i] = feedback
            ws[feedback_after] = feedback_

            counter += 1 # Right here!
        except IndexError:
            print('not aligned!') # and this is correct actually too somehow. CHECK!
    wb.save('base_data.xlsx')


for company in company_list:
    scrape_site(company)
